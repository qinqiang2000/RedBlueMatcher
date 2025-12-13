#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
结果输出模块
统一管理匹配结果的输出格式、文件命名和路径
"""

import csv
import os
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import List, TYPE_CHECKING

if TYPE_CHECKING:
    from red_blue_matcher import MatchResult, SKUSummary, FailedMatch


@dataclass
class OutputConfig:
    """输出配置"""
    base_name: str = 'match_results'    # 基础文件名
    format: str = 'csv'                  # 输出格式: 'csv' | 'xlsx'
    add_timestamp: bool = True           # 是否添加时间戳
    output_dir: str = './output'         # 输出目录
    sheet_name: str = '匹配结果'          # xlsx sheet名称


class ResultWriter:
    """统一输出接口"""

    # CSV/XLSX 表头定义（集中管理）
    HEADERS = [
        '序号',
        '待红冲 SKU 编码',
        '该 SKU 红冲对应蓝票的fid',
        '该 SKU 红冲对应蓝票的发票号码',
        '该 SKU 红冲对应蓝票的开票日期',
        '该 SKU 红冲对应蓝票的发票行号',
        '该 SKU红冲对应蓝票行的剩余可红冲金额',
        '该 SKU红冲对应蓝票行的可红冲单价',
        '本次红冲扣除的红冲金额（正数）',
        '本次红冲扣除 SKU数量',
        '扣除本次红冲后，对应蓝票行的剩余可红冲金额',
        '是否属于整行红冲'
    ]

    def __init__(self, config: OutputConfig = None):
        self.config = config or OutputConfig()

    def build_filepath(self) -> str:
        """
        构建输出文件路径（集中管理命名规则）

        Returns:
            完整的文件路径
        """
        # 确保输出目录存在
        output_dir = self.config.output_dir
        if not os.path.isabs(output_dir):
            # 相对路径转为绝对路径（相对于项目根目录）
            base_dir = os.path.dirname(os.path.abspath(__file__))
            output_dir = os.path.join(base_dir, output_dir)
        os.makedirs(output_dir, exist_ok=True)

        # 构建文件名
        base_name = os.path.splitext(self.config.base_name)[0]  # 去掉可能的扩展名

        if self.config.add_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"{base_name}_{timestamp}"
        else:
            filename = base_name

        # 添加扩展名
        ext = '.xlsx' if self.config.format == 'xlsx' else '.csv'
        filename = filename + ext

        return os.path.join(output_dir, filename)

    def write(self,
              results: List['MatchResult'],
              sku_summaries: List['SKUSummary'] = None,
              failed_matches: List['FailedMatch'] = None) -> str:
        """
        写入匹配结果

        Args:
            results: 匹配结果列表
            sku_summaries: SKU统计汇总列表
            failed_matches: 匹配失败记录列表

        Returns:
            实际写入的文件路径
        """
        filepath = self.build_filepath()

        if self.config.format == 'xlsx':
            self._write_xlsx(results, filepath, sku_summaries, failed_matches)
        else:
            self._write_csv(results, filepath)

        return filepath

    def _result_to_row(self, r: 'MatchResult') -> list:
        """
        将单个 MatchResult 转换为输出行

        Args:
            r: 匹配结果对象

        Returns:
            输出行数据列表
        """
        # 本次红冲扣除 SKU数量 (保留10位小数)
        red_quantity = (r.matched_amount / r.unit_price).quantize(
            Decimal('0.0000000001'), ROUND_HALF_UP
        )

        # 扣除本次红冲后，对应蓝票行的剩余可红冲金额
        remaining_after = r.remain_amount_before - r.matched_amount

        # 是否属于整行红冲 (剩余金额在0到0.10元之间)
        is_full_line_red = '是' if (Decimal('0') <= remaining_after <= Decimal('0.10')) else '否'

        # 格式化开票日期
        issue_date = r.fissuetime.strftime('%Y-%m-%d') if r.fissuetime else ''

        return [
            r.seq,                                    # 序号
            r.sku_code,                               # 待红冲 SKU 编码
            r.blue_fid,                               # 该 SKU 红冲对应蓝票的fid
            r.blue_invoice_no,                        # 该 SKU 红冲对应蓝票的发票号码
            issue_date,                               # 该 SKU 红冲对应蓝票的开票日期
            r.blue_entryid,                           # 该 SKU 红冲对应蓝票的发票行号
            f"{r.remain_amount_before:.2f}",          # 该 SKU红冲对应蓝票行的剩余可红冲金额
            f"{r.unit_price:.10f}",                   # 该 SKU红冲对应蓝票行的可红冲单价
            f"{r.matched_amount:.2f}",                # 本次红冲扣除的红冲金额（正数）
            f"{red_quantity:.10f}",                   # 本次红冲扣除 SKU数量（10位小数）
            f"{remaining_after:.2f}",                 # 扣除本次红冲后，对应蓝票行的剩余可红冲金额
            is_full_line_red                          # 是否属于整行红冲
        ]

    def _summary_to_row(self, s: 'SKUSummary') -> list:
        """
        将单个 SKUSummary 转换为输出行

        Args:
            s: SKU统计汇总对象

        Returns:
            输出行数据列表
        """
        return [
            s.seq,                                      # 序号
            s.sku_code,                                 # 待红冲 SKU 编码
            f"{s.original_total_amount:.2f}",           # 待红冲 SKU 总金额
            f"{s.original_total_quantity:.10f}",        # 待红冲 SKU 总数量
            f"{s.original_avg_price:.10f}",             # 待红冲 SKU 平均单价
            s.matched_blue_count,                       # 该 SKU红冲扣除蓝票的总数量（票数）
            f"{s.matched_total_amount:.2f}",            # 该 SKU 红冲扣除蓝票的总金额
            f"{s.matched_total_quantity:.10f}",         # 该 SKU 红冲扣除蓝票的总数量（计算）
            s.matched_line_count,                       # 该 SKU红冲扣除蓝票的总行数
            f"{s.remaining_blue_amount:.2f}"            # 该 SKU红冲扣除蓝票上，剩余可红冲金额合计
        ]

    def _failed_to_row(self, f: 'FailedMatch') -> list:
        """
        将单个 FailedMatch 转换为输出行

        Args:
            f: 匹配失败记录对象

        Returns:
            输出行数据列表
        """
        return [
            f.seq,                                      # 序号
            f.negative_fid,                             # 负数单据fid
            f.negative_entryid,                         # 负数单据行号
            f.negative_billno,                          # 负数单据编号
            f.sku_code,                                 # SKU编码
            f.goods_name,                               # 商品名称
            f.tax_rate,                                 # 税率
            f"{f.amount:.2f}",                          # 金额（负数）
            f"{f.quantity:.10f}",                       # 数量（负数）
            f"{f.tax:.2f}",                             # 税额（负数）
            f.failed_reason                             # 失败原因
        ]

    def _write_csv(self, results: List['MatchResult'], filepath: str):
        """CSV输出"""
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            writer.writerow(self.HEADERS)
            for r in results:
                writer.writerow(self._result_to_row(r))

    def _write_xlsx(self,
                    results: List['MatchResult'],
                    filepath: str,
                    sku_summaries: List['SKUSummary'] = None,
                    failed_matches: List['FailedMatch'] = None):
        """
        XLSX输出（支持多个sheet）

        依赖: openpyxl
        安装: pip install openpyxl
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            raise ImportError(
                "导出xlsx格式需要安装openpyxl: pip install openpyxl"
            )

        wb = Workbook()

        # Sheet 1: SKU 红冲扣除蓝票明细表
        ws1 = wb.active
        ws1.title = self.config.sheet_name

        # 写入表头
        ws1.append(self.HEADERS)
        # 表头加粗
        for cell in ws1[1]:
            cell.font = Font(bold=True)

        # 写入数据
        for r in results:
            ws1.append(self._result_to_row(r))

        # 将大整数列格式化为文本（防止科学计数法显示）
        # 列索引: C=3(fid), D=4(发票号码), F=6(发票行号)
        text_columns = [3, 4, 6]  # C, D, F 列
        for row in range(2, ws1.max_row + 1):  # 从第2行开始（跳过表头）
            for col in text_columns:
                cell = ws1.cell(row=row, column=col)
                # 将值转换为字符串，添加前缀单引号强制为文本格式
                if cell.value is not None:
                    cell.value = str(cell.value)
                    cell.number_format = '@'  # 设置为文本格式

        # Sheet 2: SKU 统计汇总表
        if sku_summaries:
            ws2 = wb.create_sheet(title='SKU 统计汇总表')
            self._write_summary_sheet(ws2, sku_summaries)

        # Sheet 3: 匹配失败记录表
        if failed_matches:
            ws3 = wb.create_sheet(title='匹配失败记录表')
            self._write_failed_sheet(ws3, failed_matches)

        wb.save(filepath)

    def _write_summary_sheet(self, ws, summaries: List['SKUSummary']):
        """写入 SKU 统计汇总表"""
        from openpyxl.styles import Font

        # 表头
        headers = [
            '序号',
            '待红冲 SKU 编码',
            '待红冲 SKU 总金额',
            '待红冲 SKU 总数量',
            '待红冲 SKU 平均单价',
            '该 SKU红冲扣除蓝票的总数量',
            '该 SKU 红冲扣除蓝票的总金额',
            '该 SKU 红冲扣除蓝票的总数量（按红冲扣除金额/可红冲单价，计算出来的数量）',
            '该 SKU红冲扣除蓝票的总行数',
            '该 SKU红冲扣除蓝票上，对应蓝票行剩余可红冲金额的合计总金额'
        ]

        ws.append(headers)
        # 表头加粗
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 写入数据
        for s in summaries:
            ws.append(self._summary_to_row(s))

    def _write_failed_sheet(self, ws, failed_matches: List['FailedMatch']):
        """写入匹配失败记录表"""
        from openpyxl.styles import Font

        # 表头
        headers = [
            '序号',
            '负数单据fid',
            '负数单据行号',
            '负数单据编号',
            'SKU编码',
            '商品名称',
            '税率',
            '金额（负数）',
            '数量（负数）',
            '税额（负数）',
            '失败原因'
        ]

        ws.append(headers)
        # 表头加粗
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # 写入数据
        for f in failed_matches:
            ws.append(self._failed_to_row(f))

        # 将大整数列格式化为文本（防止科学计数法显示）
        # 列索引: B=2(fid), C=3(行号)
        text_columns = [2, 3]  # B, C 列
        for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
            for col in text_columns:
                cell = ws.cell(row=row, column=col)
                # 将值转换为字符串，添加前缀单引号强制为文本格式
                if cell.value is not None:
                    cell.value = str(cell.value)
                    cell.number_format = '@'  # 设置为文本格式
