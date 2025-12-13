#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
统计红票数量工具
根据匹配结果分析需要开具的红票数量

规则: 一张红票只能对应一张蓝票
     因此需要开具的红票数量 = 被红冲的蓝票数量（即整票红冲判断表的行数）
"""

import openpyxl
import sys
from pathlib import Path


def count_red_invoices(excel_file: str):
    """
    统计需要开具的红票数量

    Args:
        excel_file: 匹配结果Excel文件路径
    """
    print(f"\n正在分析文件: {excel_file}")
    print("=" * 60)

    try:
        # 加载Excel文件
        wb = openpyxl.load_workbook(excel_file, read_only=True)

        # 读取Sheet 1: SKU 红冲扣除蓝票明细表
        sheet_name = 'SKU 红冲扣除蓝票明细表'
        if sheet_name not in wb.sheetnames:
            print(f"❌ 错误: 文件中未找到'{sheet_name}'")
            print(f"   可用的sheet: {wb.sheetnames}")
            return

        ws = wb[sheet_name]

        # 统计D列（蓝票发票号码）的唯一值
        # D列索引为3（从0开始）
        invoice_numbers = set()

        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
            invoice_no = row[0]
            if invoice_no:  # 过滤空值
                invoice_numbers.add(invoice_no)

        unique_count = len(invoice_numbers)
        total_rows = ws.max_row - 1  # 总数据行数（去掉表头）

        print(f"\n📊 统计结果:")
        print(f"   明细表总行数: {total_rows} 行")
        print(f"   唯一蓝票发票号码数: {unique_count} 张")
        print(f"   需要开具的红票数量: {unique_count} 张")
        print(f"\n说明:")
        print(f"   - 一张红票只能对应一张蓝票")
        print(f"   - D列（蓝票发票号码）的唯一值 = 需要开具的红票数")
        print(f"   - 相当于Excel公式: =ROWS(UNIQUE(D2:D{total_rows + 1}))")

        # 显示前10张不同的发票号码
        print(f"\n📋 发票号码样例（前10张）:")
        for i, inv_no in enumerate(sorted(invoice_numbers)[:10], start=1):
            print(f"   {i}. {inv_no}")

        if unique_count > 10:
            print(f"   ... (还有 {unique_count - 10} 张)")

        wb.close()

        print("\n" + "=" * 60)
        print(f"✅ 结论: 需要开具 {unique_count} 张红票")
        print("=" * 60)

    except FileNotFoundError:
        print(f"❌ 错误: 文件不存在: {excel_file}")
    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback
        traceback.print_exc()


def main():
    """主函数"""
    if len(sys.argv) < 2:
        # 自动查找最新的输出文件
        output_dir = Path('./output')
        if not output_dir.exists():
            print("❌ 错误: output目录不存在")
            print("\n用法:")
            print("  python count_red_invoices.py <Excel文件路径>")
            print("\n示例:")
            print("  python count_red_invoices.py ./output/match_results_20251213_113609.xlsx")
            return

        # 查找最新的xlsx文件
        xlsx_files = sorted(output_dir.glob('match_results_*.xlsx'),
                           key=lambda x: x.stat().st_mtime,
                           reverse=True)

        if not xlsx_files:
            print("❌ 错误: output目录下没有找到匹配结果文件")
            print("\n用法:")
            print("  python count_red_invoices.py <Excel文件路径>")
            return

        excel_file = str(xlsx_files[0])
        print(f"ℹ️  自动选择最新文件: {excel_file}")
    else:
        excel_file = sys.argv[1]

    count_red_invoices(excel_file)


if __name__ == '__main__':
    main()
